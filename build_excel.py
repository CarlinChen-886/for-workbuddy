#!/usr/bin/env python3
# 电商礼盒优惠对比 → Excel 生成器（多 Sheet 版：首Sheet「计算逻辑与规则」→ 品牌Sheet）
# 输入: data.json (anchors_specs + products[])，每个 product 可含 brand/upload_date/price_check/platform_coupons
# 输出: 1个规则Sheet + N个品牌Sheet，含图片嵌入
import json, os, sys, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

def pct(x):
    if x is None:
        return "—"
    return f"{x*100:.1f}%"


def _safe_str(s):
    """防止以 = + - @ 开头的字符串被 Excel/WPS 误判为公式 → 显示 #NAME?

    OCR 偶尔会把 ¥ 前面带噪声或把产品名前缀识别成 `=`/`+`/`-`/`@`，
    openpyxl 把以这些符号开头的字符串当作公式写入，打开时 WPS 报 #NAME?。
    修正：行首出现的公式触发符替换成全角等价（仍可读），不改变正文。
    """
    if not isinstance(s, str) or not s:
        return s
    if "\r" in s:
        s = s.replace("\r", "")
    repl = {"=": "＝", "+": "＋", "-": "－", "@": "＠"}
    if s[0] in repl:
        s = repl[s[0]] + s[1:]
    return s

# 图片显示尺寸（像素）→ 对应 Excel 列宽/行高
IMG_TARGET_H = 300          # 显示高度
IMG_COL_W = 34              # A 列宽度（约 238px）
IMG_ROW_H = 225             # 行高（约 300px @96dpi）
DATE_COL_W = 14             # B 列（上传日期/周期）宽度


def _prorate_platform(plat_coupons, official):
    """平台券门槛折算：主品价不足门槛时按比例折算（规则12）"""
    total = 0.0
    result = []
    for pc in plat_coupons:
        amount = pc.get("amount", 0)
        condition = pc.get("condition", "")
        threshold = None
        m = re.search(r'满[¥￥]?\s*(\d{3,5})', condition)
        if m:
            threshold = float(m.group(1))
        if threshold and official < threshold:
            effective = amount * (official / threshold)
            total += effective
            note = f"需凑单·按比例折算: ¥{amount}×(¥{official}/¥{threshold:g})=¥{effective:.0f}"
            result.append({**pc, "effective_amount": effective, "prorate_note": note, "prorated": True})
        else:
            total += amount
            result.append({**pc, "effective_amount": amount, "prorate_note": "", "prorated": False})
    return result, total


def _add_logic_sheet(wb, anchors_specs):
    """创建独立的「计算逻辑与规则」Sheet（首位）"""
    ws = wb.create_sheet("计算逻辑与规则", 0)
    ws.column_dimensions["A"].width = 120
    title = ws.cell(row=1, column=1, value="计算逻辑与规则")
    title.font = Font(bold=True, size=14, color="305496")
    logic = ["一、正装官方售价锚点（天猫品牌官方旗舰店在售价；多规格取最高容量常规规格单 ml 价，不用平均）"]
    if anchors_specs:
        for k, v in anchors_specs.items():
            logic.append(f"  {k}：{v}")
    logic += [
        "",
        "二、硬规则（用户确认）",
        "  1) 价格唯一来源：天猫品牌官方旗舰店（规则 1/6）",
        "  2) 多规格估值：选长期售卖常规规格中容量最高者的单 ml 价，排除限定/特殊超大装（规则 1/2）",
        "  3) 赠品价值 = 赠品 ml 数 × 正装单 ml 官方售价（按最高容量常规规格单 ml 价）（规则 3/4）",
        "  4) 美妆工具（非官方店在售，如按摩棒）→ ¥0（规则 3）",
        "  5) 赠品必须读全文字：主图 logo、文案、注释、细则中写出的赠品均计入（规则 5）",
        "  6) 主图/人工补充才可录入优惠，禁止根据大促臆测（规则 6）",
        "  7) 商家优惠：店铺会员券、直播/自播间货补（限时直降/补贴/立减等）→「商家优惠层」(L/M/N)（规则 7）",
        "  8) 平台券：美妆加补券、88VIP消费券/专享券等→「平台券层」(O/P/Q)，另算（规则 8）",
        "  9) 合计/至高叠减只作分项校验，禁止重复计算（规则 9）",
        "  10) 购物金/储值权益→L列单独展示，完全不参与折扣计算（规则 10）",
        "  11) 两年对比：去年价≠当前价 → 采用校验出的过去年价（规则 11）",
        "  12) 平台券门槛折算：主品价不足门槛时按比例折算（规则 12）",
        "",
        "三、折扣率 / off% 口径",
        "  · 仅赠品 off% = 赠品总价值 / (主品正价 + 赠品总价值)",
        "  · 赠品含金量 = 赠品总价值 / 到手价",
        "  【商家优惠层】（规则 7）",
        "  · 含赠品含商家优惠 off% = 1 − 商家到手价 / (主品正价 + 赠品总价值)",
        "  【平台券层】（规则 8/9/12）",
        "  · 叠加平台券后到手价 = 商家到手价 − 平台券折算后合计",
        "  · 主品价不足门槛时，按比例：实际抵扣 = 券面额 × (主品价/门槛)",
        "  【购物金/储值权益】（规则 10）",
        "  · 完全不计入任何 off%/折扣率",
        "",
        "四、总价值 = 主品正价 + 赠品总价值",
        "五、公式示例：面霜 60ml¥2455（最高容量常规规格）→ 单 ml 价 = ¥40.92/ml（规则 1/2/4）",
    ]
    for i, line in enumerate(logic, start=2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=11)
        if line and not line.startswith("  ") and ("、" in line or line[0] in "一二三四五"):
            cell.font = Font(bold=True, size=12, color="305496")


def _fill_sheet(ws, products, anchors_specs):
    """填充单 Sheet：表头 + 数据行 + 计算逻辑同页；products 与 anchors_specs 已为该 Sheet 范围"""
    # 列 A = 原图；列 B = 上传日期/周期；C~Q = 数据
    # 商家优惠层（直播间货补等）: L/M/N；平台优惠层（88VIP消费券/美妆品类券等）: O/P/Q
    headers = ["原图", "上传日期/周期", "商品", "主品", "主品内构成", "主品活动周期/Logo文案",
               "主品官方售价(¥)", "赠品明细", "赠品对应价值(¥)", "赠品总价值(¥)",
               "仅赠品 off%/含金量", "商家优惠罗列(直播间货补/金额)", "商家优惠后到手价(¥)", "含赠品含商家优惠 off%/折扣率",
               "平台券罗列(88VIP/品类券/惊喜券等)", "叠加平台券后到手价(¥)", "叠加平台券后 off%/折扣率"]
    ws.append(headers)
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="305496")
    for c in ws[1]:
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, p in enumerate(products, start=1):
        gifts = p.get("gifts", [])
        gift_total = p.get("gift_total_value")
        if gift_total is None:
            gift_total = sum(g.get("value", 0) for g in gifts)
        official = p["official_price"]
        final = p["final_price"]

        # 主品官方售价列展示（含可选的去年价校验信息）
        price_check = p.get("price_check")
        if price_check:
            pc_cur = price_check.get("current")
            pc_past = price_check.get("past_price")
            pc_year = price_check.get("past_year", "")
            pc_note = price_check.get("note", "")
            if pc_cur is not None and pc_past is not None and pc_cur == pc_past:
                tag = "一致·无变动·正常锚定"
            else:
                used_past = (price_check.get("used") == pc_past) if "used" in price_check else (pc_past != pc_cur)
                tag = "采用去年价" if used_past else "采用当前价"
            official_disp = f"¥{official:,.0f}\n[去年价校验] 当前¥{pc_cur} / {pc_year}年¥{pc_past} → {tag}"
            if pc_note:
                official_disp += f"\n注：{pc_note}"
        else:
            official_disp = official
        total_value = official + gift_total

        gift_off = (gift_total / total_value) if total_value else 0
        gift_ratio = (gift_total / final) if final else 0
        total_off = (1 - final / total_value) if total_value else 0
        disc_rate = (final / total_value) if total_value else 0

        gift_names = "\n".join(_safe_str(g["name"]) for g in gifts) or "无"
        gift_vals = "\n".join(_safe_str(f'{g["name"]}：¥{g.get("value",0):.2f}') for g in gifts) or "无"
        promos = p.get("promotions", [])
        promo_txt = "\n".join(_safe_str(f'{x["type"]}：¥{x.get("amount",0)}（{x.get("condition","")}）') for x in promos) or "无"
        # 购物金/储值权益（规则 10）：单独展示，不参与折扣计算
        sc = p.get("shopping_credits")
        if sc and sc.get("amount", 0) > 0:
            promo_txt += f'\n\n【储值权益 · 不计入折扣】\n{_safe_str(sc.get("type","购物金"))}：¥{sc.get("amount",0):g}（{sc.get("condition","")}）'
            if sc.get("note"):
                promo_txt += f'\n注：{sc.get("note")}'
        elif sc and sc.get("type"):
            promo_txt += f'\n\n【储值权益 · 不计入折扣】\n{_safe_str(sc.get("type","购物金"))}：待确认（{sc.get("condition","")}）'

        # 平台优惠层（规则 8/12：门槛不足时按比例折算）
        plat_coupons = p.get("platform_coupons", [])
        plat_coupons, plat_total = _prorate_platform(plat_coupons, official)
        final_plat = final - plat_total
        if plat_coupons:
            plat_lines = []
            for x in plat_coupons:
                base = _safe_str(f'{x.get("name","平台券")}：¥{x.get("amount",0)}（{x.get("condition","")}）')
                if x.get("prorated"):
                    base += f'\n  {x.get("prorate_note","")}'
                plat_lines.append(base)
            plat_txt = "\n".join(plat_lines)
            plat_txt += f"\n平台券折算后合计：¥{plat_total:.0f}"
            plat_strength = (plat_total / final) if final else 0
            total_off_plat = (1 - final_plat / total_value) if total_value else 0
            disc_rate_plat = (final_plat / total_value) if total_value else 0
            price_off_plat = (1 - final_plat / official) if official else 0
            plat_off_disp = (
                f"平台券力度 {pct(plat_strength)}（省¥{plat_total}，相对商家到手价）\n"
                f"叠加平台券价格off% {pct(price_off_plat)}（对主品正价）\n"
                f"含赠品综合off% {pct(total_off_plat)}\n折扣率 {pct(disc_rate_plat)}（约{disc_rate_plat*10:.1f}折）"
            )
        else:
            plat_txt = "无（非大促/无平台券）"
            plat_off_disp = "—（无平台券，同商家优惠层）"

        row = [
            "",  # A 留空，图片随后锚定
            p.get("upload_date", ""),
            _safe_str(p.get("name", "")),
            _safe_str(p.get("main_product", "")),
            _safe_str(p.get("main_composition", "")),
            _safe_str(f'周期：{p.get("period","")}\n文案：{p.get("logo_text","")}'),
            official_disp,
            gift_names,
            gift_vals,
            round(gift_total, 2),
            f"赠品off% {pct(gift_off)}（占整体）\n赠品含金量 {pct(gift_ratio)}（赠品/到手价）",
            promo_txt,
            final,
            f"off% {pct(total_off)}\n折扣率 {pct(disc_rate)}（约{disc_rate*10:.1f}折）",
            plat_txt,
            final_plat,
            plat_off_disp,
        ]  # fmt: skip
        # 对 row 里所有字符串值统一防公式触发（_safe_str 已对单字段处理过gift/promo/plat，对其他字段再扫一次保险）
        row = [_safe_str(v) if isinstance(v, str) else v for v in row]
        ws.append(row)
        r = ws.max_row
        for c in ws[r]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border

        img_path = p.get("image_path", "")
        if img_path and os.path.exists(img_path):
            img = XLImage(img_path)
            scale = IMG_TARGET_H / img.height if img.height else 1
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            img.anchor = f"A{r}"
            ws.add_image(img)
        else:
            ws.cell(row=r, column=1, value="（图片缺失）")

    # 列宽（17 列）
    widths = [IMG_COL_W, DATE_COL_W, 16, 18, 16, 32, 12, 22, 26, 11, 20, 30, 13, 24, 32, 14, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 44
    for r in range(2, 2 + len(products)):
        ws.row_dimensions[r].height = IMG_ROW_H
    ws.freeze_panes = "C2"


def build_workbook(data):
    """首 Sheet「计算逻辑与规则」，其余按 brand 自动分组"""
    products = data.get("products", [])
    anchors_specs = data.get("anchors_specs", {})
    if isinstance(anchors_specs, list):
        anchors_specs = {f"锚点{i+1}": v for i, v in enumerate(anchors_specs) if v}

    wb = Workbook()
    wb.remove(wb.active)

    # 首位：独立的「计算逻辑与规则」Sheet
    _add_logic_sheet(wb, anchors_specs)

    # 按 brand 分组
    groups = {}
    order = []
    for p in products:
        b = p.get("brand") or "默认"
        if b not in groups:
            groups[b] = []
            order.append(b)
        groups[b].append(p)

    for brand in order:
        ws = wb.create_sheet(brand)
        _fill_sheet(ws, groups[brand], anchors_specs)

    return wb

def build(data_path, out_path):
    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)
    wb = build_workbook(data)
    wb.save(out_path)
    brands = [p.get("brand", "默认") for p in data.get("products", [])]
    n_brands = len(set(brands)) if brands else 0
    n_sheets = n_brands + 1  # +1 for 计算逻辑与规则
    print(f"SAVED: {out_path}  (Sheet数: {n_sheets} = 1规则 + {n_brands}品牌, 商品数: {len(products := data.get('products', []))})")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: build_excel.py <data.json> <out.xlsx>")
        sys.exit(1)
    build(sys.argv[1], sys.argv[2])